import base64
import io
from openpyxl import load_workbook
from odoo import models, fields, api


class WizardSendEmail(models.TransientModel):
    _name = "wizard.send.email"
    _description = "Wizard for sending emails"

    upload_file = fields.Binary(string="Upload Excel File", required=True)
    filename = fields.Char(string="File Name")

    def send_emails(self):
        if not self.upload_file:
            raise UserError("Please upload an Excel file before sending emails.")

        # Step 1: Decode uploaded file
        file_content = base64.b64decode(self.upload_file)
        workbook = load_workbook(filename=io.BytesIO(file_content))
        sheet = workbook.active

        # Step 2: Parse rows and send emails
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            email, subject = row
            if email and subject:
                self.env["mail.mail"].create(
                    {
                        "subject": subject,
                        "body_html": "<p>Welcome in GymBeam!</p>",
                        "email_to": email,
                        "email_from": self.env.user.email or "noreply@example.com",
                    }
                ).send()

        return {"type": "ir.actions.act_window_close"}
